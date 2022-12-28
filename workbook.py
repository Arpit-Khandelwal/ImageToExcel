import pytesseract
import openpyxl


pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'


# Load the image and perform OCR on it
image = "1.jpeg"
text = pytesseract.image_to_string(image)

# Split the OCR text into lines
lines = text.split("\n")

# Create a new workbook and add a sheet to it
workbook = openpyxl.Workbook()
sheet = workbook.active

# Write the extracted information to the sheet
headers = []
data = []
for i, line in enumerate(lines):
  if i == 0:
    # Parse the headers from the first line
    headers = line.split()
  else:
    # Parse the data from the remaining lines
    data.append(line.split())

# Write the headers to the sheet
for i, header in enumerate(headers):
  sheet.cell(row=1, column=i+1).value = header

# Write the data to the sheet
for i, row in enumerate(data):
  for j, cell in enumerate(row):
    sheet.cell(row=i+2, column=j+1).value = cell

# Save the workbook to a file
workbook.save("output.xlsx")